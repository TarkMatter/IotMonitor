"""
InfluxDB から月次センサーデータを取得して Excel レポートを自動生成するモジュール。

起動方法（当月レポートを生成）:
    python -m reporter.main

    または python コードから:
        from reporter.main import generate_monthly_report
        path = generate_monthly_report(2026, 6)
"""

from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from collector.influx_client import EnvInfluxClient
from common import messages

# --- スタイル定数 ---
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ROW_EVEN_FILL = PatternFill("solid", fgColor="EBF5FF")
_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# データシートのカラムヘッダー
_DATA_HEADERS = ["日時", "デバイスID", "温度(℃)", "湿度(%)", "CO₂(ppm)", "気圧(hPa)"]

# サマリーシートの行ラベル
_SUMMARY_LABELS = ["温度(℃)", "湿度(%)", "CO₂(ppm)"]


def _apply_header_style(cell: openpyxl.cell.Cell) -> None:
    """セルにヘッダースタイル（背景色・フォント・中央揃え・罫線）を適用する。

    Args:
        cell: スタイルを適用する openpyxl セル
    """
    cell.fill      = _HEADER_FILL
    cell.font      = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    cell.border    = _THIN_BORDER


def _write_data_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, data: list[dict]) -> None:
    """環境データシートにヘッダーとデータ行を書き込む。

    偶数行にうっすらした青の背景を付けて視認性を高める。

    Args:
        ws:   書き込み先ワークシート
        data: EnvInfluxClient.query_recent() が返す辞書リスト
    """
    # ヘッダー行
    for col, header in enumerate(_DATA_HEADERS, start=1):
        _apply_header_style(ws.cell(row=1, column=col, value=header))

    # データ行
    for row_idx, record in enumerate(data, start=2):
        values = [
            record["time"],
            record["device_id"],
            record["temperature"],
            record["humidity"],
            record["co2"],
            record["pressure"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = _ROW_EVEN_FILL

    # カラム幅の自動調整（最大 50 文字まで）
    for col_cells in ws.columns:
        max_length = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 2, 50)


def _write_summary_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, data: list[dict]) -> None:
    """サマリーシートに平均・最大・最小の集計表と折れ線グラフを書き込む。

    Args:
        ws:   書き込み先ワークシート
        data: EnvInfluxClient.query_recent() が返す辞書リスト
    """
    # 集計ヘッダー
    for col, header in enumerate(["指標", "平均", "最大", "最小"], start=1):
        _apply_header_style(ws.cell(row=1, column=col, value=header))

    if not data:
        return

    # 各センサー値のリストを作成（None を除外）
    temps  = [d["temperature"] for d in data if d["temperature"] is not None]
    humids = [d["humidity"]    for d in data if d["humidity"]    is not None]
    co2s   = [d["co2"]        for d in data if d["co2"]         is not None]

    summary_rows = [
        (
            "温度(℃)",
            round(sum(temps) / len(temps), 1) if temps else None,
            max(temps) if temps else None,
            min(temps) if temps else None,
        ),
        (
            "湿度(%)",
            round(sum(humids) / len(humids), 1) if humids else None,
            max(humids) if humids else None,
            min(humids) if humids else None,
        ),
        (
            "CO₂(ppm)",
            round(sum(co2s) / len(co2s)) if co2s else None,
            max(co2s) if co2s else None,
            min(co2s) if co2s else None,
        ),
    ]

    for row_idx, row_data in enumerate(summary_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER

    # 温度の折れ線グラフ（直近 100 件を使用してグラフを見やすくする）
    sample_count = min(len(data), 100)
    temp_values  = [d["temperature"] for d in data[-sample_count:]]

    # グラフ用の一時データをシート右側に書き込む
    chart_start_col = 6
    ws.cell(row=1, column=chart_start_col, value="温度推移（サンプル）")
    for i, val in enumerate(temp_values, start=2):
        ws.cell(row=i, column=chart_start_col, value=val)

    chart = LineChart()
    chart.title  = "温度推移"
    chart.y_axis.title = "温度(℃)"
    chart.style  = 10
    chart.width  = 20
    chart.height = 12

    data_ref = Reference(ws, min_col=chart_start_col, min_row=2, max_row=sample_count + 1)
    chart.add_data(data_ref)
    ws.add_chart(chart, "A6")


def generate_monthly_report(
    year: int,
    month: int,
    device_id: str | None = None,
    output_dir: str = ".",
) -> Path:
    """指定年月の環境データを InfluxDB から取得して Excel レポートを生成する。

    Args:
        year:       レポート対象年
        month:      レポート対象月（1〜12）
        device_id:  特定デバイスに絞り込む場合に指定。None で全デバイス
        output_dir: Excel ファイルの出力先ディレクトリパス

    Returns:
        生成した Excel ファイルの Path オブジェクト

    Raises:
        ValueError: month が 1〜12 の範囲外の場合
    """
    if not 1 <= month <= 12:
        raise ValueError(f"month は 1〜12 の整数で指定してください: {month}")

    # 対象月全体を InfluxDB に対してクエリするために時間数を算出する
    start = date(year, month, 1)
    end   = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    hours = int((end - start).days * 24)

    influx = EnvInfluxClient()
    data   = influx.query_recent(hours=hours, device_id=device_id)

    wb = openpyxl.Workbook()
    ws_data    = wb.active
    ws_data.title = "環境データ"
    ws_summary = wb.create_sheet("サマリー")

    _write_data_sheet(ws_data, data)
    _write_summary_sheet(ws_summary, data)

    fname = f"env_report_{year}{month:02d}.xlsx"
    out   = Path(output_dir) / fname
    wb.save(out)
    print(messages.REPORT_GENERATED.format(path=out))
    return out


if __name__ == "__main__":
    now = datetime.now()
    generate_monthly_report(now.year, now.month)
