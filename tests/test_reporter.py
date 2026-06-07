"""
reporter/main.py の単体テスト。

InfluxDB への接続はモックし、Excel ファイルの生成・シート構造・
値の正確さを一時ディレクトリで実際に検証する。
"""

import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

# EnvInfluxClient をモックしてインポート時の接続を防ぐ
_MOCK_INFLUX = MagicMock()
with patch("collector.influx_client.EnvInfluxClient", return_value=_MOCK_INFLUX):
    from reporter.main import generate_monthly_report

# テスト用のダミーセンサーデータ（1時間分 × 6レコード）
_SAMPLE_DATA = [
    {
        "time":        f"2026-06-07T0{i}:00:00+00:00",
        "device_id":   "raspi-01",
        "temperature": 20.0 + i,
        "humidity":    50.0 + i,
        "co2":         600 + i * 10,
        "pressure":    1013.0,
    }
    for i in range(6)
]


@pytest.fixture
def tmp_dir():
    """テスト用の一時ディレクトリを提供する。"""
    with tempfile.TemporaryDirectory() as d:
        yield d


@pytest.fixture(autouse=True)
def patch_influx():
    """全テストで EnvInfluxClient をモックする。"""
    with patch("reporter.main.EnvInfluxClient", return_value=_MOCK_INFLUX):
        _MOCK_INFLUX.query_recent.return_value = _SAMPLE_DATA
        yield _MOCK_INFLUX


class TestGenerateMonthlyReport:
    """generate_monthly_report() の基本動作テスト。"""

    def test_returns_path_object(self, tmp_dir):
        result = generate_monthly_report(2026, 6, output_dir=tmp_dir)
        assert isinstance(result, Path)

    def test_file_is_created(self, tmp_dir):
        path = generate_monthly_report(2026, 6, output_dir=tmp_dir)
        assert path.exists()

    def test_filename_format(self, tmp_dir):
        """ファイル名が env_report_YYYYMM.xlsx 形式になる。"""
        path = generate_monthly_report(2026, 6, output_dir=tmp_dir)
        assert path.name == "env_report_202606.xlsx"

    def test_december_filename(self, tmp_dir):
        """12月のファイル名が正しい（年またぎ計算の確認）。"""
        path = generate_monthly_report(2026, 12, output_dir=tmp_dir)
        assert path.name == "env_report_202612.xlsx"

    def test_invalid_month_raises(self, tmp_dir):
        with pytest.raises(ValueError):
            generate_monthly_report(2026, 13, output_dir=tmp_dir)

    def test_invalid_month_zero_raises(self, tmp_dir):
        with pytest.raises(ValueError):
            generate_monthly_report(2026, 0, output_dir=tmp_dir)

    def test_device_id_passed_to_influx(self, tmp_dir, patch_influx):
        generate_monthly_report(2026, 6, device_id="raspi-01", output_dir=tmp_dir)
        call_kwargs = patch_influx.query_recent.call_args
        assert call_kwargs.kwargs.get("device_id") == "raspi-01"


class TestExcelSheetStructure:
    """生成された Excel ファイルのシート構造テスト。"""

    @pytest.fixture
    def workbook(self, tmp_dir):
        path = generate_monthly_report(2026, 6, output_dir=tmp_dir)
        return openpyxl.load_workbook(path)

    def test_has_two_sheets(self, workbook):
        assert len(workbook.sheetnames) == 2

    def test_sheet_names(self, workbook):
        assert workbook.sheetnames == ["環境データ", "サマリー"]

    def test_data_sheet_header_row(self, workbook):
        ws = workbook["環境データ"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["日時", "デバイスID", "温度(℃)", "湿度(%)", "CO₂(ppm)", "気圧(hPa)"]

    def test_data_sheet_row_count(self, workbook):
        """ヘッダー1行 + データ6行 = 7行になる。"""
        ws = workbook["環境データ"]
        assert ws.max_row == 1 + len(_SAMPLE_DATA)

    def test_temperature_values_written_correctly(self, workbook):
        ws = workbook["環境データ"]
        temps = [ws.cell(row=r, column=3).value for r in range(2, 2 + len(_SAMPLE_DATA))]
        expected = [d["temperature"] for d in _SAMPLE_DATA]
        assert temps == expected

    def test_summary_sheet_has_header(self, workbook):
        ws = workbook["サマリー"]
        assert ws.cell(row=1, column=1).value == "指標"

    def test_summary_average_temperature(self, workbook):
        """サマリーシートの平均温度が正しく計算されている。"""
        ws = workbook["サマリー"]
        expected_avg = round(sum(d["temperature"] for d in _SAMPLE_DATA) / len(_SAMPLE_DATA), 1)
        assert ws.cell(row=2, column=2).value == expected_avg


class TestEmptyData:
    """データが空の場合でもファイル生成が正常に完了することを確認する。"""

    def test_empty_data_creates_file(self, tmp_dir, patch_influx):
        patch_influx.query_recent.return_value = []
        path = generate_monthly_report(2026, 6, output_dir=tmp_dir)
        assert path.exists()

    def test_empty_data_has_header_only(self, tmp_dir, patch_influx):
        patch_influx.query_recent.return_value = []
        path = generate_monthly_report(2026, 6, output_dir=tmp_dir)
        wb = openpyxl.load_workbook(path)
        # ヘッダー行のみで、データ行なし
        assert wb["環境データ"].max_row == 1
